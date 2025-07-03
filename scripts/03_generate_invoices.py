# scripts/03_generate_invoices.py

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

TEMPLATE = Path("templates/invoice_template.xlsx")
DATA_FP = Path("data/processed/sales_clean.csv")
OUT_DIR = Path("data/processed/invoices")

def generate_all():
    df = pd.read_csv(DATA_FP, parse_dates=["ContractDate"])
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for _, row in df.iterrows():
        wb = load_workbook(TEMPLATE)
        ws = wb.active
        # Assume cell B2 holds customer name, B3 date, B4 amount, etc.
        ws["B2"] = row["CustomerName"]
        ws["B3"] = row["ContractDate"].strftime("%Y-%m-%d")
        ws["B4"] = row["TotalAmount"]
        # Mark invoice as generated
        filename = OUT_DIR / f"Invoice_{row.ContractID}.xlsx"
        wb.save(filename)
        print("Generated", filename)

if __name__ == "__main__":
    generate_all()
